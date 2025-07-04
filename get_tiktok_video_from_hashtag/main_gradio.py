import gradio as gr
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
from tqdm import tqdm
import re
import pandas as pd
from datetime import datetime
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
import matplotlib.pyplot as plt
import tempfile
import os
from openpyxl import load_workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, Alignment, PatternFill
import threading
from collections import deque

# Global variables
driver = None

# API Key Management System
class APIKeyManager:
    def __init__(self, api_keys):
        self.api_keys = deque(api_keys)
        self.current_key = None
        self.failed_keys = set()
        self.lock = threading.Lock()
        self._get_next_key()
    
    def _get_next_key(self):
        """Lấy key tiếp theo từ queue"""
        with self.lock:
            if len(self.api_keys) == 0:
                # Reset queue nếu đã hết key
                self.api_keys = deque([k for k in List_API_KEY if k not in self.failed_keys])
                if len(self.api_keys) == 0:
                    # Nếu tất cả key đều failed, reset lại toàn bộ
                    self.failed_keys.clear()
                    self.api_keys = deque(List_API_KEY)
            
            if self.api_keys:
                self.current_key = self.api_keys.popleft()
                print(f"🔑 Đang sử dụng API Key: {self.current_key[:8]}...")
            return self.current_key
    
    def get_current_key(self):
        """Lấy key hiện tại"""
        return self.current_key
    
    def mark_key_failed(self, error_msg=""):
        """Đánh dấu key hiện tại là failed và chuyển sang key tiếp theo"""
        with self.lock:
            if self.current_key:
                self.failed_keys.add(self.current_key)
                print(f"❌ API Key {self.current_key[:8]}... đã failed: {error_msg}")
                self._get_next_key()
                print(f"🔄 Chuyển sang API Key mới: {self.current_key[:8]}...")
    
    def is_quota_error(self, error):
        """Kiểm tra xem có phải lỗi quota không"""
        if isinstance(error, HttpError):
            error_details = str(error)
            quota_indicators = [
                "quotaExceeded", 
                "dailyLimitExceeded",
                "rateLimitExceeded",
                "quota exceeded",
                "daily limit exceeded",
                "rate limit exceeded"
            ]
            return any(indicator.lower() in error_details.lower() for indicator in quota_indicators)
        return False
    
    def execute_with_retry(self, func, *args, **kwargs):
        """Thực thi function với retry khi gặp lỗi quota"""
        max_retries = len(List_API_KEY)
        
        for attempt in range(max_retries):
            try:
                current_key = self.get_current_key()
                if not current_key:
                    raise Exception("Không còn API key khả dụng")
                
                # Thêm api_key vào kwargs
                kwargs['api_key'] = current_key
                result = func(*args, **kwargs)
                return result
                
            except Exception as e:
                error_msg = str(e)
                
                # Kiểm tra nếu là lỗi quota hoặc key bị chặn
                if (self.is_quota_error(e) or 
                    "forbidden" in error_msg.lower() or
                    "invalid" in error_msg.lower() or
                    "disabled" in error_msg.lower()):
                    
                    self.mark_key_failed(error_msg)
                    
                    if attempt == max_retries - 1:
                        raise Exception(f"Tất cả API key đều failed. Lỗi cuối: {error_msg}")
                    
                    print(f"⚠️ Đang thử lại với API key mới... (lần {attempt + 1}/{max_retries})")
                    time.sleep(1)  # Delay một chút trước khi retry
                    continue
                else:
                    # Lỗi khác không liên quan đến API key
                    raise e
        
        raise Exception("Đã hết số lần thử")

# Khởi tạo API Key Manager global
List_API_KEY = [
    'AIzaSyBkbdni528jYjh4Igj5GAvHDV6q4hrR6Kk',
    'AIzaSyBvtOpz_xQTxaCYJGMyT9OIKtu-nlAPdvw',
    'AIzaSyC1EOQa2qAief4m6tWmWCygeqRDmdUKMUQ',
    'AIzaSyBgnX69psjLuqYvX5U_J8MnY2-PEHOLs3c',
    'AIzaSyBLI0ABhHbkTVE0x9ZgDMxwZhwkHuLp6Ec'
]

api_manager = APIKeyManager(List_API_KEY)

def format_number(num):
    """Format số thành dạng K, M, B"""
    try:
        num = int(num)
        if num >= 1_000_000_000:
            return f"{num/1_000_000_000:.1f}B"
        elif num >= 1_000_000:
            return f"{num/1_000_000:.1f}M"
        elif num >= 1_000:
            return f"{num/1_000:.1f}K"
        else:
            return str(num)
    except:
        return str(num)

def open_browser():
    """Mở trình duyệt Chrome"""
    global driver
    try:
        if driver is not None:
            try:
                driver.quit()
            except:
                pass
        
        options = Options()
        options.add_argument("--window-size=1200,800")
        options.add_argument("--disable-extensions")
        options.add_argument("--disable-plugins")
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option('useAutomationExtension', False)
        
        # Để browser hiển thị (không headless)
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
        driver.get("https://www.youtube.com")
        
        return "✅ Đã mở trình duyệt Chrome thành công!"
    except Exception as e:
        return f"❌ Lỗi khi mở trình duyệt: {str(e)}"

def close_browser():
    """Đóng trình duyệt"""
    global driver
    try:
        if driver is not None:
            driver.quit()
            driver = None
            return "✅ Đã đóng trình duyệt!"
        else:
            return "⚠️ Chưa có trình duyệt nào được mở!"
    except Exception as e:
        return f"❌ Lỗi khi đóng trình duyệt: {str(e)}"

def get_channels_urls(hashtag, limit, driver):
    """Lấy danh sách URLs của các kênh từ hashtag"""
    try:
        # 1. Mở YouTube
        driver.get("https://www.youtube.com")
        time.sleep(2)

        # 2. Tìm kiếm hashtag
        search_box = driver.find_element(By.NAME, "search_query")
        search_box.clear()
        search_box.send_keys(f"{hashtag}")
        search_box.send_keys(Keys.RETURN)
        time.sleep(3)

        # 3. Chuyển sang tab "Shorts" (nếu có)
        try:
            shorts_tab = driver.find_element(By.XPATH, "//yt-chip-cloud-chip-renderer//span[contains(text(), 'Shorts')]")
            shorts_tab.click()
            time.sleep(2)
        except Exception:
            pass

        # 4. Lấy kênh cho tới khi đủ LIMIT
        channels = set()
        last_height = driver.execute_script("return document.documentElement.scrollHeight")
        MAX_SCROLL_ATTEMPTS = 5
        scroll_attempts = 0
        
        while len(channels) < limit and scroll_attempts < MAX_SCROLL_ATTEMPTS:
            shorts = driver.find_elements(By.XPATH, "//a[contains(@href, '/shorts/')]")
            for short in shorts:
                if len(channels) >= limit:
                    break
                try:
                    parent = short.find_element(By.XPATH, "./../../..")
                    selectors = [
                        ".//a[contains(@href, '/@')]",
                        ".//a[contains(@href, '/channel/')]", 
                        ".//a[contains(@href, '/c/')]"
                    ]

                    for selector in selectors:
                        try:
                            channel_elem = parent.find_element(By.XPATH, selector)
                            channel_name = channel_elem.text
                            channel_url = channel_elem.get_attribute("href")
                            channels.add((channel_name, channel_url))
                            break
                        except:
                            continue
                except Exception:
                    continue

            if len(channels) >= limit:
                break

            # Cuộn xuống
            driver.execute_script("window.scrollTo(0, document.documentElement.scrollHeight);")
            time.sleep(0.5)  
            
            new_height = driver.execute_script("return document.documentElement.scrollHeight")
            if new_height == last_height:
                scroll_attempts += 1
                time.sleep(1)
            else:
                scroll_attempts = 0
            last_height = new_height

        return [url for name, url in list(channels)]
    except Exception as e:
        print(f"Lỗi khi lấy kênh: {e}")
        return []

def get_channel_info_from_url(url, api_key):
    """Lấy thông tin kênh trực tiếp từ URL YouTube"""
    youtube = build('youtube', 'v3', developerKey=api_key)
    
    # Direct channel ID
    channel_id_match = re.search(r'youtube\.com/channel/([A-Za-z0-9_-]+)', url)
    if channel_id_match:
        channel_id = channel_id_match.group(1)
        request = youtube.channels().list(
            part='snippet,statistics',
            id=channel_id,
            fields='items(id,snippet/title,snippet/publishedAt,statistics/subscriberCount,statistics/videoCount,statistics/viewCount)'
        )
        response = request.execute()
        return response.get('items', [])
    
    # Username format
    handle_match = re.search(r'youtube\.com/@([A-Za-z0-9._-]+)', url)
    if handle_match:
        handle = handle_match.group(1)
        request = youtube.channels().list(
            part='snippet,statistics',
            forHandle=handle,
            fields='items(id,snippet/title,snippet/publishedAt,statistics/subscriberCount,statistics/videoCount,statistics/viewCount)'
        )
        response = request.execute()
        return response.get('items', [])
    
    # Legacy username format
    username_match = re.search(r'youtube\.com/(?:user/|c/)([A-Za-z0-9_-]+)', url)
    if username_match:
        username = username_match.group(1)
        request = youtube.channels().list(
            part='snippet,statistics',
            forUsername=username,
            fields='items(id,snippet/title,snippet/publishedAt,statistics/subscriberCount,statistics/videoCount,statistics/viewCount)'
        )
        response = request.execute()
        return response.get('items', [])
    
    return []

def extract_channel_id_from_url(url):
    """Trích xuất channel_id từ URL"""
    channel_id_match = re.search(r'youtube\.com/channel/([A-Za-z0-9_-]+)', url)
    if channel_id_match:
        return channel_id_match.group(1)
    return None

def get_channels_info_by_ids(channel_ids, api_key):
    """Lấy thông tin nhiều kênh bằng channel_id"""
    youtube = build('youtube', 'v3', developerKey=api_key)
    request = youtube.channels().list(
        part='snippet,statistics',
        id=','.join(channel_ids),
        fields='items(id,snippet/title,snippet/publishedAt,statistics/subscriberCount,statistics/videoCount,statistics/viewCount)'
    )
    response = request.execute()
    return response.get('items', [])

def get_multiple_channels_from_urls_batch(channel_urls):
    """Lấy thông tin nhiều kênh từ danh sách URLs với API key rotation"""
    global api_manager
    
    all_data = []
    id_url_map = {}
    ids = []
    fallback_urls = []

    # Tách channel_id nếu có thể
    for url in channel_urls:
        cid = extract_channel_id_from_url(url)
        if cid:
            ids.append(cid)
            id_url_map[cid] = url
        else:
            fallback_urls.append(url)

    # Chia batch 50 id/lần với retry mechanism
    for i in range(0, len(ids), 50):
        batch_ids = ids[i:i+50]
        try:
            # Sử dụng API manager để execute với retry
            items = api_manager.execute_with_retry(get_channels_info_by_ids, batch_ids)
            for item in items:
                all_data.append({
                    'id': item['id'],
                    'title': item['snippet']['title'],
                    'publishedAt': item['snippet']['publishedAt'],
                    'subscriberCount': item['statistics'].get('subscriberCount', 0),
                    'videoCount': item['statistics'].get('videoCount', 0),
                    'viewCount': item['statistics'].get('viewCount', 0),
                    'source_url': id_url_map[item['id']]
                })
        except Exception as e:
            print(f"Lỗi khi lấy batch ids: {e}")

    # Với các url không có channel_id
    for url in fallback_urls:
        try:
            # Sử dụng API manager để execute với retry  
            channel_items = api_manager.execute_with_retry(get_channel_info_from_url, url)
            for item in channel_items:
                all_data.append({
                    'id': item['id'],
                    'title': item['snippet']['title'],
                    'publishedAt': item['snippet']['publishedAt'],
                    'subscriberCount': item['statistics'].get('subscriberCount', 0),
                    'videoCount': item['statistics'].get('videoCount', 0),
                    'viewCount': item['statistics'].get('viewCount', 0),
                    'source_url': url
                })
        except Exception as e:
            print(f"Lỗi khi lấy thông tin từ {url}: {e}")

    return pd.DataFrame(all_data)

def plot_subscribers_chart(df):
    """Vẽ biểu đồ subscribers"""
    if df.empty:
        return None
    
    # Sắp xếp theo subscribers giảm dần
    df_sorted = df.sort_values('subscriberCount', ascending=False)
    
    plt.style.use('seaborn-v0_8')
    fig, ax = plt.subplots(figsize=(12, 8))
    
    subscribers = [int(x) for x in df_sorted['subscriberCount']]
    titles = [title[:20] + '...' if len(title) > 20 else title for title in df_sorted['title']]
    
    bars = ax.bar(titles, subscribers, color=['#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4', '#FECA57', 
                                           '#FF9FF3', '#54A0FF', '#5F27CD', '#FD79A8', '#FDCB6E'] * (len(titles)//10+1))
    
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: format_number(x)))
    ax.set_ylabel('Subscribers', fontsize=12, fontweight='bold')
    ax.set_xlabel('Channel', fontsize=12, fontweight='bold')
    ax.set_title('Channels by Subscribers', fontsize=14, fontweight='bold', pad=20)
    plt.xticks(rotation=45, ha='right')
    
    for bar, sub in zip(bars, subscribers):
        height = bar.get_height()
        ax.text(bar.get_x() + bar.get_width()/2., height + height*0.01,
                format_number(sub), ha='center', va='bottom', fontweight='bold')
    
    ax.grid(True, alpha=0.3, axis='y')
    plt.tight_layout()
    
    return fig

def save_to_excel_with_chart(df, hashtag):
    """Lưu Excel với biểu đồ"""
    if df.empty:
        return None
    
    # Tạo file tạm thời
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    file_path = temp_file.name
    temp_file.close()
    
    # Chuẩn bị data
    df_export = df.copy()
    
    # Convert to numeric
    df_export['subscriberCount'] = pd.to_numeric(df_export['subscriberCount'], errors='coerce').fillna(0)
    df_export['videoCount'] = pd.to_numeric(df_export['videoCount'], errors='coerce').fillna(0)
    df_export['viewCount'] = pd.to_numeric(df_export['viewCount'], errors='coerce').fillna(0)
    
    # Tạo cột formatted
    df_export['subscriberCount_formatted'] = df_export['subscriberCount'].apply(format_number)
    df_export['videoCount_formatted'] = df_export['videoCount'].apply(format_number)
    df_export['viewCount_formatted'] = df_export['viewCount'].apply(format_number)
    
    # Sắp xếp
    df_export = df_export.sort_values('subscriberCount', ascending=False)
    
    # Sắp xếp columns
    df_export = df_export[['title', 'subscriberCount_formatted', 'videoCount_formatted', 
                          'viewCount_formatted', 'publishedAt', 'subscriberCount', 
                          'videoCount', 'viewCount', 'source_url']]
    
    # Lưu DataFrame
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df_export.to_excel(writer, index=False, sheet_name='Channels')
    
    # Thêm biểu đồ vào file Excel
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Format header
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Format numeric columns
    for row in range(2, ws.max_row + 1):
        ws[f'F{row}'].number_format = '#,##0'
        ws[f'G{row}'].number_format = '#,##0'
        ws[f'H{row}'].number_format = '#,##0'
    
    # Auto-fit columns
    column_widths = {
        'A': 30, 'B': 15, 'C': 15, 'D': 15, 'E': 12,
        'F': 15, 'G': 15, 'H': 15, 'I': 40
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Tạo biểu đồ
    max_row = ws.max_row
    chart_start_row = ws.max_row + 3
    
    chart = BarChart()
    chart.title = "All Channels by Subscribers"
    chart.y_axis.title = 'Subscribers'
    chart.x_axis.title = 'Channel'
    chart.width = 25
    chart.height = 15
    chart.style = 10
    
    data = Reference(ws, min_col=6, min_row=1, max_row=max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=max_row)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend = None
    
    ws.add_chart(chart, f"A{chart_start_row}")
    
    # Thêm title cho section chart
    ws[f'A{chart_start_row-2}'] = "📊 BIỂU ĐỒ SO SÁNH SUBSCRIBERS"
    ws[f'A{chart_start_row-2}'].font = Font(size=14, bold=True, color="366092")
    
    # Đặt hyperlink cho cột I
    for row in range(2, ws.max_row + 1):
        cell = ws[f'I{row}']
        url = cell.value
        if url:
            cell.hyperlink = url
            cell.style = "Hyperlink"
    
    wb.save(file_path)
    return file_path

def search_channels(hashtag, limit, progress=gr.Progress()):
    """Tìm kiếm kênh YouTube"""
    global driver
    
    if driver is None:
        gr.Error("Vui lòng mở trình duyệt trước!")
        return None, None, None, "❌ Vui lòng mở trình duyệt trước!"
    
    if not hashtag:
        gr.Error("Vui lòng nhập hashtag!")
        return None, None, None, "❌ Vui lòng nhập hashtag!"
    
    progress(0.1, desc="Đang tìm kiếm kênh...")
    
    try:
        # Lấy danh sách kênh
        channel_urls = get_channels_urls(hashtag, int(limit), driver)
        
        if not channel_urls:
            gr.Warning("Không tìm thấy kênh nào!")
            return None, None, None, "⚠️ Không tìm thấy kênh nào!"
        
        progress(0.5, desc=f"Đã tìm thấy {len(channel_urls)} kênh, đang lấy thông tin...")
        
        # Lấy thông tin chi tiết với API key rotation
        df = get_multiple_channels_from_urls_batch(channel_urls)
        
        if df.empty:
            gr.Warning("Không lấy được thông tin kênh!")
            return None, None, None, "⚠️ Không lấy được thông tin kênh!"
        
        progress(0.8, desc="Đang tạo biểu đồ và file Excel...")
        
        # Format DataFrame cho hiển thị
        df_display = df.copy()
        df_display['subscriberCount'] = df_display['subscriberCount'].apply(format_number)
        df_display['videoCount'] = df_display['videoCount'].apply(format_number)
        df_display['viewCount'] = df_display['viewCount'].apply(format_number)
        df_display['publishedAt'] = pd.to_datetime(df_display['publishedAt']).dt.strftime('%Y-%m-%d')
        
        # Đổi tên cột cho dễ đọc
        df_display = df_display.rename(columns={
            'title': 'Tên kênh',
            'subscriberCount': 'Subscribers',
            'videoCount': 'Videos',
            'viewCount': 'Views',
            'publishedAt': 'Ngày tạo',
            'source_url': 'Link'
        })
        
        # Sắp xếp columns
        df_display = df_display[['Tên kênh', 'Subscribers', 'Videos', 'Views', 'Ngày tạo', 'Link']]
        
        # Vẽ biểu đồ
        fig = plot_subscribers_chart(df)
        
        # Lưu file Excel
        excel_path = save_to_excel_with_chart(df, hashtag)
        
        progress(1.0, desc="Hoàn thành!")
        
        # Hiển thị thông tin API key được sử dụng
        current_key = api_manager.get_current_key()
        status_msg = f"✅ Đã tìm thấy {len(df)} kênh! (API Key: {current_key[:8]}...)"
        
        return df_display, fig, excel_path, status_msg
        
    except Exception as e:
        gr.Error(f"Lỗi: {str(e)}")
        return None, None, None, f"❌ Lỗi: {str(e)}"

# Tạo giao diện Gradio
with gr.Blocks(title="YouTube Channel Finder by Hashtag", theme=gr.themes.Soft()) as demo:
    gr.Markdown(
        """
        # 🎬 YouTube Channel Finder by Hashtag
        
        Tìm kiếm các kênh YouTube theo hashtag và xuất dữ liệu ra file Excel.
        """
    )
    
    with gr.Row():
        with gr.Column(scale=1):
            gr.Markdown("### 🔧 Điều khiển")
            
            # Browser controls
            with gr.Group():
                gr.Markdown("**Trình duyệt Chrome**")
                browser_status = gr.Textbox(label="Trạng thái", value="Chưa mở trình duyệt", interactive=False)
                with gr.Row():
                    open_btn = gr.Button("🌐 Mở trình duyệt", variant="primary")
                    close_btn = gr.Button("❌ Đóng trình duyệt", variant="secondary")
            
            # Search controls
            with gr.Group():
                gr.Markdown("**Tìm kiếm**")
                hashtag_input = gr.Textbox(label="Hashtag", placeholder="Nhập hashtag (VD: #food, #music)")
                limit_input = gr.Number(label="Số lượng kênh", value=20, minimum=1, maximum=100, step=1)
                search_btn = gr.Button("🔍 Tìm kiếm", variant="primary", size="lg")
            
            # Status
            status_text = gr.Textbox(label="Kết quả", value="", interactive=False)
        
        with gr.Column(scale=3):
            gr.Markdown("### 📊 Kết quả")
            
            # Results tabs
            with gr.Tabs():
                with gr.TabItem("📋 Bảng dữ liệu"):
                    results_df = gr.DataFrame(
                        headers=["Tên kênh", "Subscribers", "Videos", "Views", "Ngày tạo", "Link"],
                        datatype=["str", "str", "str", "str", "str", "str"],
                        interactive=False,
                        wrap=True
                    )
                
                with gr.TabItem("📈 Biểu đồ"):
                    chart_plot = gr.Plot()
            
            # Download button
            with gr.Row():
                excel_file = gr.File(label="📥 Tải file Excel", visible=False)
    
    # Event handlers
    def update_browser_status(msg):
        return msg
    
    open_btn.click(
        fn=lambda: update_browser_status(open_browser()),
        outputs=[browser_status]
    )
    
    close_btn.click(
        fn=lambda: update_browser_status(close_browser()),
        outputs=[browser_status]
    )
    
    search_btn.click(
        fn=search_channels,
        inputs=[hashtag_input, limit_input],
        outputs=[results_df, chart_plot, excel_file, status_text]
    ).then(
        fn=lambda x: gr.update(visible=x is not None),
        inputs=[excel_file],
        outputs=[excel_file]
    )
    
    # Examples
    gr.Examples(
        examples=[
            ["#food", 20],
            ["#music", 30],
            ["#travel", 25],
            ["#gaming", 15],
            ["#football", 20]
        ],
        inputs=[hashtag_input, limit_input],
        label="Ví dụ"
    )

# Chạy ứng dụng
if __name__ == "__main__":
    demo.launch(
        share=False,
        server_name="0.0.0.0",
        server_port=7860,
        inbrowser=True
    ) 