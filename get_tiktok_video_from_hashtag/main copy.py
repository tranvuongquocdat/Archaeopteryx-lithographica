import tkinter as tk
from tkinter import ttk, messagebox
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import time
from tqdm import tqdm
import re
import pandas as pd
from datetime import datetime
from googleapiclient.discovery import build
import pandas as pd  # Thêm import pandas
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from tkinter import filedialog
import webbrowser

def format_number(num):
    """Format số thành dạng K, M, B"""
    try:
        num = int(num)
        if num >= 1_000_000_000:
            return f"{num/1_000_000_000:.1f}B"
        elif num >= 1_000_000:
            return f"{num/1_000_000:.1f}M"
        elif num >= 1_000:
            return f"{num/1_000:.1f}K"
        else:
            return str(num)
    except:
        return str(num)

def get_channels_urls(hashtag, limit, driver):

    try:
        # 1. Mở YouTube
        driver.get("https://www.youtube.com")

        # 2. Tìm kiếm hashtag
        search_box = driver.find_element(By.NAME, "search_query")
        search_box.send_keys(f"{hashtag}")
        search_box.send_keys(Keys.RETURN)
        time.sleep(3)

        # 3. Chuyển sang tab "Shorts" (nếu có)
        try:
            shorts_tab = driver.find_element(By.XPATH, "//yt-chip-cloud-chip-renderer//span[contains(text(), 'Shorts')]")
            shorts_tab.click()
            time.sleep(2)
        except Exception:
            pass  # Nếu không có tab Shorts thì bỏ qua

        # 4. Lấy kênh cho tới khi đủ LIMIT
        channels = set()
        pbar = tqdm(total=limit, desc="Số kênh đã lấy", ncols=80)
        last_height = driver.execute_script("return document.documentElement.scrollHeight")
        MAX_SCROLL_ATTEMPTS = 5
        scroll_attempts = 0
        while len(channels) < limit and scroll_attempts < MAX_SCROLL_ATTEMPTS:
            shorts = driver.find_elements(By.XPATH, "//a[contains(@href, '/shorts/')]")
            for short in shorts:
                if len(channels) >= limit:
                    break
                try:
                    parent = short.find_element(By.XPATH, "./../../..")
                    # Thử nhiều selector khác nhau
                    selectors = [
                        ".//a[contains(@href, '/@')]",
                        ".//a[contains(@href, '/channel/')]", 
                        ".//a[contains(@href, '/c/')]"
                    ]

                    for selector in selectors:
                        try:
                            channel_elem = parent.find_element(By.XPATH, selector)
                            channel_name = channel_elem.text
                            channel_url = channel_elem.get_attribute("href")
                            before = len(channels)
                            channels.add((channel_name, channel_url))
                            after = len(channels)
                            if after > before:
                                pbar.update(1)
                            break
                        except:
                            continue
                except Exception as e:
                    print(f"Lỗi khi lấy thông tin kênh: {e}")  # Debug
                    continue

            if len(channels) >= limit:
                break

            # Cuộn xuống và đợi lâu hơn
            driver.execute_script("window.scrollTo(0, document.documentElement.scrollHeight);")
            time.sleep(0.5)  
            
            new_height = driver.execute_script("return document.documentElement.scrollHeight")
            if new_height == last_height:
                scroll_attempts += 1
                time.sleep(1)  # Đợi thêm trước khi thử lại
            else:
                scroll_attempts = 0  # Reset nếu có nội dung mới
            last_height = new_height
        pbar.close()

        # 5. return ra danh sách kênh
        print(f"Đã lấy {len(channels)} kênh:")
        return [url for name, url in list(channels)]
    except Exception as e:
        print(f"Lỗi khi lấy kênh: {e}")
        return []
    


    """
    Lấy thông tin nhiều kênh từ danh sách URLs
    """
    all_data = []
    
    for url in channel_urls:
        try:
            channel_items = get_channel_info_from_url(url, api_key)
            for item in channel_items:
                all_data.append({
                    'id': item['id'],
                    'title': item['snippet']['title'],
                    'publishedAt': item['snippet']['publishedAt'],
                    'subscriberCount': item['statistics'].get('subscriberCount', 0),
                    'videoCount': item['statistics'].get('videoCount', 0),
                    'viewCount': item['statistics'].get('viewCount', 0),
                    'source_url': url  # Thêm URL gốc để tham khảo
                })
        except Exception as e:
            print(f"Lỗi khi lấy thông tin từ {url}: {e}")
            continue
    
    return pd.DataFrame(all_data)

def get_channel_info_from_url(url, api_key):
    """
    Lấy thông tin kênh trực tiếp từ URL YouTube
    """
    youtube = build('youtube', 'v3', developerKey=api_key)
    
    # Direct channel ID (youtube.com/channel/UCxxxxxx)
    channel_id_match = re.search(r'youtube\.com/channel/([A-Za-z0-9_-]+)', url)
    if channel_id_match:
        channel_id = channel_id_match.group(1)
        request = youtube.channels().list(
            part='snippet,statistics',
            id=channel_id,
            fields='items(id,snippet/title,snippet/publishedAt,statistics/subscriberCount,statistics/videoCount,statistics/viewCount)'
        )
        response = request.execute()
        return response.get('items', [])
    
    # Username format (youtube.com/@username)
    handle_match = re.search(r'youtube\.com/@([A-Za-z0-9._-]+)', url)
    if handle_match:
        handle = handle_match.group(1)
        request = youtube.channels().list(
            part='snippet,statistics',
            forHandle=handle,  # Không cần @ prefix
            fields='items(id,snippet/title,snippet/publishedAt,statistics/subscriberCount,statistics/videoCount,statistics/viewCount)'
        )
        response = request.execute()
        return response.get('items', [])
    
    # Legacy username format (youtube.com/user/username hoặc youtube.com/c/customname)
    username_match = re.search(r'youtube\.com/(?:user/|c/)([A-Za-z0-9_-]+)', url)
    if username_match:
        username = username_match.group(1)
        request = youtube.channels().list(
            part='snippet,statistics',
            forUsername=username,
            fields='items(id,snippet/title,snippet/publishedAt,statistics/subscriberCount,statistics/videoCount,statistics/viewCount)'
        )
        response = request.execute()
        return response.get('items', [])
    
    return []

def extract_channel_id_from_url(url):
    """
    Trích xuất channel_id từ URL nếu có thể, trả về None nếu không tìm được.
    """
    # Direct channel ID (youtube.com/channel/UCxxxxxx)
    channel_id_match = re.search(r'youtube\.com/channel/([A-Za-z0-9_-]+)', url)
    if channel_id_match:
        return channel_id_match.group(1)
    return None

def get_channels_info_by_ids(channel_ids, api_key):
    """
    Lấy thông tin nhiều kênh bằng channel_id (tối đa 50/lần)
    """
    youtube = build('youtube', 'v3', developerKey=api_key)
    request = youtube.channels().list(
        part='snippet,statistics',
        id=','.join(channel_ids),
        fields='items(id,snippet/title,snippet/publishedAt,statistics/subscriberCount,statistics/videoCount,statistics/viewCount)'
    )
    response = request.execute()
    return response.get('items', [])

def get_multiple_channels_from_urls_batch(channel_urls, api_key):
    """
    Lấy thông tin nhiều kênh từ danh sách URLs, tối ưu quota bằng batch channel_id
    """
    all_data = []
    id_url_map = {}
    ids = []
    fallback_urls = []

    # Tách channel_id nếu có thể
    for url in channel_urls:
        cid = extract_channel_id_from_url(url)
        if cid:
            ids.append(cid)
            id_url_map[cid] = url
        else:
            fallback_urls.append(url)

    # Chia batch 50 id/lần
    for i in range(0, len(ids), 50):
        batch_ids = ids[i:i+50]
        try:
            items = get_channels_info_by_ids(batch_ids, api_key)
            for item in items:
                all_data.append({
                    'id': item['id'],
                    'title': item['snippet']['title'],
                    'publishedAt': item['snippet']['publishedAt'],
                    'subscriberCount': item['statistics'].get('subscriberCount', 0),
                    'videoCount': item['statistics'].get('videoCount', 0),
                    'viewCount': item['statistics'].get('viewCount', 0),
                    'source_url': id_url_map[item['id']]
                })
        except Exception as e:
            print(f"Lỗi khi lấy batch ids {batch_ids}: {e}")

    # Với các url không có channel_id, fallback sang hàm cũ (tốn quota hơn)
    for url in fallback_urls:
        print(f"Channel url không có channel_id: {url}")
        try:
            channel_items = get_channel_info_from_url(url, api_key)
            for item in channel_items:
                all_data.append({
                    'id': item['id'],
                    'title': item['snippet']['title'],
                    'publishedAt': item['snippet']['publishedAt'],
                    'subscriberCount': item['statistics'].get('subscriberCount', 0),
                    'videoCount': item['statistics'].get('videoCount', 0),
                    'viewCount': item['statistics'].get('viewCount', 0),
                    'source_url': url
                })
        except Exception as e:
            print(f"Lỗi khi lấy thông tin từ {url}: {e}")

    return pd.DataFrame(all_data)

def chunks(lst, n):
    for i in range(0, len(lst), n):
        yield lst[i:i + n]

# --- Khởi tạo Chrome với webdriver-manager (optimized) ---
options = Options()
options.add_argument("--window-size=800,600")  # Nhỏ hơn
options.add_argument("--disable-extensions")
options.add_argument("--disable-plugins")
options.add_argument("--disable-images")  # Tắt load ảnh để nhanh hơn
options.add_argument("--disable-javascript")  # Tắt một số JS không cần thiết
options.add_experimental_option("excludeSwitches", ["enable-automation"])
options.add_experimental_option('useAutomationExtension', False)
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
driver.get("https://www.youtube.com")  # Chỉ mở sẵn YouTube
API_KEY = 'AIzaSyBkbdni528jYjh4Igj5GAvHDV6q4hrR6Kk'
list_API_KEY = []

def plot_subscribers_bar(df, frame):
    """Vẽ biểu đồ subscribers cho toàn bộ kênh"""
    for widget in frame.winfo_children():
        widget.destroy()
    
    # Sắp xếp theo subscribers giảm dần, KHÔNG lấy top 10 nữa
    df_sorted = df.sort_values('subscriberCount', ascending=False)
    
    plt.style.use('seaborn-v0_8')
    fig, ax = plt.subplots(figsize=(12, 6))
    
    subscribers = [int(x) for x in df_sorted['subscriberCount']]
    titles = [title[:20] + '...' if len(title) > 20 else title for title in df_sorted['title']]
    
    bars = ax.bar(titles, subscribers, color=['#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4', '#FECA57', 
                                           '#FF9FF3', '#54A0FF', '#5F27CD', '#FD79A8', '#FDCB6E'] * (len(titles)//10+1))
    
    ax.yaxis.set_major_formatter(plt.FuncFormatter(lambda x, p: format_number(x)))
    ax.set_ylabel('Subscribers', fontsize=12, fontweight='bold')
    ax.set_xlabel('Channel', fontsize=12, fontweight='bold')
    ax.set_title('Channels by Subscribers', fontsize=14, fontweight='bold', pad=20)
    plt.xticks(rotation=45, ha='right')
    
    for bar, sub in zip(bars, subscribers):
        height = bar.get_height()
        ax.text(bar.get_x() + bar.get_width()/2., height + height*0.01,
                format_number(sub), ha='center', va='bottom', fontweight='bold')
    
    ax.grid(True, alpha=0.3, axis='y')
    plt.tight_layout()
    
    canvas = FigureCanvasTkAgg(fig, master=frame)
    canvas.draw()
    canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

def save_to_excel_with_chart(df, default_filename):
    """Lưu Excel với biểu đồ đẹp hơn và đặt ở cuối"""
    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx", 
        initialfile=default_filename, 
        filetypes=[("Excel files", "*.xlsx")]
    )
    if not file_path:
        return
    
    # Chuẩn bị data và đảm bảo số liệu là numeric
    df_export = df.copy()
    
    # Convert to numeric để đảm bảo biểu đồ hoạt động
    df_export['subscriberCount'] = pd.to_numeric(df_export['subscriberCount'], errors='coerce').fillna(0)
    df_export['videoCount'] = pd.to_numeric(df_export['videoCount'], errors='coerce').fillna(0)
    df_export['viewCount'] = pd.to_numeric(df_export['viewCount'], errors='coerce').fillna(0)
    
    # Tạo cột formatted
    df_export['subscriberCount_formatted'] = df_export['subscriberCount'].apply(format_number)
    df_export['videoCount_formatted'] = df_export['videoCount'].apply(format_number)
    df_export['viewCount_formatted'] = df_export['viewCount'].apply(format_number)
    
    # Sắp xếp theo subscribers giảm dần
    df_export = df_export.sort_values('subscriberCount', ascending=False)
    
    # Sắp xếp columns: formatted columns ở đầu, raw numbers ở cuối
    df_export = df_export[['title', 'subscriberCount_formatted', 'videoCount_formatted', 
                          'viewCount_formatted', 'publishedAt', 'subscriberCount', 
                          'videoCount', 'viewCount', 'source_url']]
    
    # Lưu DataFrame
    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        df_export.to_excel(writer, index=False, sheet_name='Channels')
    
    # Thêm biểu đồ vào file Excel
    from openpyxl import load_workbook
    from openpyxl.chart import BarChart, Reference
    from openpyxl.styles import Font, Alignment, PatternFill, NamedStyle, numbers
    
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Format header
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Format numeric columns (F, G, H - subscriberCount, videoCount, viewCount)
    for row in range(2, ws.max_row + 1):
        # subscriberCount column (F)
        ws[f'F{row}'].number_format = '#,##0'
        # videoCount column (G) 
        ws[f'G{row}'].number_format = '#,##0'
        # viewCount column (H)
        ws[f'H{row}'].number_format = '#,##0'
    
    # Auto-fit columns
    column_widths = {
        'A': 30,  # title
        'B': 15,  # subscriberCount_formatted
        'C': 15,  # videoCount_formatted
        'D': 15,  # viewCount_formatted
        'E': 12,  # publishedAt
        'F': 15,  # subscriberCount (raw)
        'G': 15,  # videoCount (raw)
        'H': 15,  # viewCount (raw)
        'I': 40   # source_url
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Lấy data để tạo chart (hiển thị toàn bộ kênh)
    max_row = ws.max_row  # Lấy tất cả rows thay vì giới hạn 11
    chart_start_row = ws.max_row + 3
    
    # Tạo biểu đồ
    chart = BarChart()
    chart.title = "All Channels by Subscribers"  # Thay đổi title
    chart.y_axis.title = 'Subscribers'
    chart.x_axis.title = 'Channel'
    chart.width = 25  # Tăng width để hiển thị nhiều data hơn
    chart.height = 15  # Tăng height để dễ nhìn hơn
    chart.style = 10  # Preset style
    
    # Reference đến dữ liệu (cột F là subscriberCount numeric)
    data = Reference(ws, min_col=6, min_row=1, max_row=max_row)  # Column F (subscriberCount)
    cats = Reference(ws, min_col=1, min_row=2, max_row=max_row)  # Column A (title)
    
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    
    # Tùy chỉnh chart
    chart.legend = None  # Bỏ legend vì chỉ có 1 series
    
    # Đặt chart ở cuối file
    ws.add_chart(chart, f"A{chart_start_row}")
    
    # Thêm title cho section chart
    ws[f'A{chart_start_row-2}'] = "📊 BIỂU ĐỒ SO SÁNH SUBSCRIBERS"
    ws[f'A{chart_start_row-2}'].font = Font(size=14, bold=True, color="366092")
    
    # Đặt hyperlink cho cột I (source_url)
    for row in range(2, ws.max_row + 1):
        cell = ws[f'I{row}']
        url = cell.value
        if url:
            cell.hyperlink = url
            cell.style = "Hyperlink"
    
    wb.save(file_path)
    print(f"Đã lưu file Excel với biểu đồ tại: {file_path}")

def run_search():
    hashtag = hashtag_var.get().strip()
    try:
        limit = int(limit_var.get())
    except:
        messagebox.showerror("Lỗi", "Số lượng kênh phải là số nguyên.")
        return
    if not hashtag:
        messagebox.showerror("Lỗi", "Vui lòng nhập hashtag.")
        return
    
    search_btn.config(state=tk.DISABLED)
    result_label.config(text="Đang tìm kiếm...")
    root.update()
    
    try:
        channel_urls = get_channels_urls(hashtag, limit, driver)
        if not channel_urls:
            messagebox.showerror("Lỗi", "Không tìm thấy kênh nào.")
            result_label.config(text="")
            search_btn.config(state=tk.NORMAL)
            return
        
        df = get_multiple_channels_from_urls_batch(channel_urls, API_KEY)
        if df.empty:
            messagebox.showerror("Lỗi", "Không lấy được thông tin kênh.")
            result_label.config(text="")
            search_btn.config(state=tk.NORMAL)
            return
        
        # Hiển thị bảng với format đẹp
        for row in tree.get_children():
            tree.delete(row)
        
        for _, row in df.iterrows():
            tree.insert('', tk.END, values=(
                row['title'], 
                format_number(row['subscriberCount']), 
                format_number(row['videoCount']), 
                format_number(row['viewCount']), 
                row['publishedAt'][:10],  # Chỉ lấy ngày
                row['source_url']         # Thêm link vào cột cuối
            ))
        
        # Vẽ biểu đồ
        plot_subscribers_bar(df, chart_frame)
        
        # Lưu lại DataFrame cho nút lưu
        root.df_result = df
        root.default_filename = f"{hashtag}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        result_label.config(text=f"Đã tìm thấy {len(df)} kênh.")
        
    except Exception as e:
        messagebox.showerror("Lỗi", str(e))
        result_label.config(text="")
    
    search_btn.config(state=tk.NORMAL)

def save_excel():
    if hasattr(root, 'df_result') and not root.df_result.empty:
        save_to_excel_with_chart(root.df_result, root.default_filename)
        messagebox.showinfo("Thành công", "Đã lưu file Excel kèm biểu đồ.")
    else:
        messagebox.showerror("Lỗi", "Chưa có dữ liệu để lưu.")

# --- UI với kích thước lớn hơn ---
root = tk.Tk()
root.title("YouTube Channel Finder by Hashtag")
root.geometry("1200x800")  # Kích thước lớn hơn
root.minsize(1000, 600)    # Kích thước tối thiểu

# Thiết lập style đẹp hơn
style = ttk.Style()
style.theme_use('clam')

input_frame = ttk.Frame(root)
input_frame.pack(padx=15, pady=10, fill=tk.X)

ttk.Label(input_frame, text="Hashtag:", font=('Arial', 10, 'bold')).pack(side=tk.LEFT)
hashtag_var = tk.StringVar()
ttk.Entry(input_frame, textvariable=hashtag_var, width=25, font=('Arial', 10)).pack(side=tk.LEFT, padx=5)

ttk.Label(input_frame, text="Số kênh:", font=('Arial', 10, 'bold')).pack(side=tk.LEFT)
limit_var = tk.StringVar(value="20")
ttk.Entry(input_frame, textvariable=limit_var, width=8, font=('Arial', 10)).pack(side=tk.LEFT, padx=5)

search_btn = ttk.Button(input_frame, text="🔍 Tìm kiếm", command=run_search)
search_btn.pack(side=tk.LEFT, padx=5)

save_btn = ttk.Button(input_frame, text="💾 Lưu Excel", command=save_excel)
save_btn.pack(side=tk.LEFT, padx=5)

result_label = ttk.Label(root, text="", font=('Arial', 10))
result_label.pack(pady=5)

# Bảng kết quả với cột rộng hơn
tree_frame = ttk.Frame(root)
tree_frame.pack(padx=15, pady=5, fill=tk.BOTH, expand=True)

tree = ttk.Treeview(tree_frame, columns=("Tên kênh", "Sub", "Video", "View", "Ngày tạo", "Link"), show="headings", height=10)
tree.heading("Tên kênh", text="Tên kênh")
tree.heading("Sub", text="Subscribers")
tree.heading("Video", text="Videos")
tree.heading("View", text="Views")
tree.heading("Ngày tạo", text="Ngày tạo")
tree.heading("Link", text="Link")  # Thêm cột Link

# Thiết lập độ rộng cột
tree.column("Tên kênh", width=300)
tree.column("Sub", width=120, anchor='center')
tree.column("Video", width=120, anchor='center')
tree.column("View", width=120, anchor='center')
tree.column("Ngày tạo", width=120, anchor='center')
tree.column("Link", width=300)  # Đặt độ rộng cho cột Link

# Scrollbar cho bảng
scrollbar = ttk.Scrollbar(tree_frame, orient="vertical", command=tree.yview)
tree.configure(yscrollcommand=scrollbar.set)

tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

# Biểu đồ với frame lớn hơn
chart_frame = ttk.LabelFrame(root, text="📊 Biểu đồ so sánh Subscribers", padding=10)
chart_frame.pack(padx=15, pady=10, fill=tk.BOTH, expand=True)

import webbrowser

def on_tree_double_click(event):
    item = tree.selection()
    if item:
        values = tree.item(item, "values")
        url = values[5]  # Cột Link là thứ 6 (index 5)
        if url:
            webbrowser.open(url)

tree.bind("<Double-1>", on_tree_double_click)

root.mainloop()
    